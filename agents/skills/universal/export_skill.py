"""
Universal Skill: Data Export (XLSX)

Exports structured project data to a multi-sheet Excel workbook:
  Sheet 1 — Project Overview
  Sheet 2 — KPIs
  Sheet 3 — Milestones
  Sheet 4 — Risks
  Sheet 5 — Budget Summary
  Sheet 6 — Action Queue

Trigger phrases: "export", "excel", "spreadsheet", "xlsx",
                 "data export", "export data", "download data"
"""

from __future__ import annotations

import logging
import os
import sqlite3
import uuid
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from agents.skills.base_skill import BaseSkill, SkillOutput
from infrastructure.project_context import ProjectContext

logger = logging.getLogger(__name__)

# Style constants
HDR_FILL  = PatternFill("solid", fgColor="1F3564")
ALT_FILL  = PatternFill("solid", fgColor="DCE6F1")
HDR_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
BODY_FONT = Font(name="Calibri", size=10)
CENT      = Alignment(horizontal="center", vertical="center")
LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN      = Side(style="thin", color="BFBFBF")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_COLORS = {
    "on_track":   "70AD47",
    "at_risk":    "FFC000",
    "off_track":  "C00000",
    "completed":  "44546A",
    "open":       "2E75B6",
    "resolved":   "70AD47",
}


def _hdr_row(ws, headers: list[str], row: int = 1) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENT
        cell.border    = BORDER


def _data_row(ws, values: list, row: int, alt: bool = False) -> None:
    fill = ALT_FILL if alt else PatternFill(fill_type=None)
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font      = BODY_FONT
        cell.fill      = fill
        cell.alignment = LEFT
        cell.border    = BORDER


def _auto_width(ws, min_w=10, max_w=45) -> None:
    for col in ws.columns:
        length = max(
            (len(str(cell.value or "")) for cell in col), default=min_w
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


class ExportSkill(BaseSkill):
    skill_id        = "data_export"
    dept_id         = "all"
    title_template  = "{project_name} — Data Export"
    description     = "Export all project data (KPIs, milestones, risks, budget, actions) to a multi-sheet Excel workbook."
    output_format   = "xlsx"
    trigger_phrases = [
        "export", "excel", "spreadsheet", "xlsx", "data export",
        "export data", "download data", "export to excel",
        "generate spreadsheet",
    ]

    async def execute(self, context: ProjectContext, params: dict[str, Any] = None) -> SkillOutput:
        params = params or {}
        title  = params.get("title") or self._make_title(context)

        try:
            db_path = getattr(context, "db_path", None)

            # ── Fetch all data ─────────────────────────────────────────────
            proj_info  = {}
            kpis       = []
            milestones = []
            risks      = []
            budget     = []
            actions    = []

            if db_path and os.path.exists(db_path):
                conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True, timeout=10)
                conn.row_factory = sqlite3.Row
                try:
                    r = conn.execute(
                        "SELECT * FROM project_details LIMIT 1"
                    ).fetchone()
                    if r:
                        proj_info = dict(r)

                    kpis = [dict(r) for r in conn.execute(
                        "SELECT name, current_value, target_value, unit, status, "
                        "trend, last_updated FROM project_kpis ORDER BY name"
                    ).fetchall()]

                    milestones = [dict(r) for r in conn.execute(
                        "SELECT name, due_date, completed_date, owner, status, description "
                        "FROM project_milestones ORDER BY due_date ASC"
                    ).fetchall()]

                    risks = [dict(r) for r in conn.execute(
                        "SELECT title, severity, impact, likelihood, status, "
                        "mitigation_plan, owner FROM project_risks ORDER BY severity DESC"
                    ).fetchall()]

                    budget = [dict(r) for r in conn.execute(
                        "SELECT category, allocated, spent, remaining, notes "
                        "FROM project_budget_lines ORDER BY category"
                    ).fetchall()]

                    actions = [dict(r) for r in conn.execute(
                        "SELECT title, category, priority, status, agent_dept, "
                        "created_at, reviewed_by FROM agent_action_queue ORDER BY created_at DESC LIMIT 50"
                    ).fetchall()]
                finally:
                    conn.close()

            # ── Build workbook ─────────────────────────────────────────────
            wb = openpyxl.Workbook()

            # Sheet 1: Overview
            ws = wb.active
            ws.title = "Overview"
            ws.row_dimensions[1].height = 30
            _hdr_row(ws, ["Field", "Value"])
            overview_fields = [
                ("Project Name",   proj_info.get("name",             context.project_id)),
                ("Tenant",         context.tenant_id),
                ("Status",         proj_info.get("status",           "—")),
                ("Health",         proj_info.get("health_status",    "—")),
                ("Priority",       proj_info.get("priority",         "—")),
                ("Start Date",     proj_info.get("start_date",       "—")),
                ("Target End",     proj_info.get("target_end_date",  "—")),
                ("Budget Total",   proj_info.get("budget_total",     0)),
                ("Budget Spent",   proj_info.get("budget_spent",     0)),
                ("Budget %",       f"{(proj_info.get('budget_spent',0)/(proj_info.get('budget_total',1) or 1)*100):.1f}%"),
                ("Description",    proj_info.get("description",      "—")),
                ("Export Date",    datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
            ]
            for i, (f, v) in enumerate(overview_fields, 2):
                _data_row(ws, [f, v], row=i, alt=(i % 2 == 0))
            _auto_width(ws)

            # Sheet 2: KPIs
            ws_kpi = wb.create_sheet("KPIs")
            _hdr_row(ws_kpi, ["KPI Name", "Current Value", "Target Value", "Unit", "Status", "Trend", "Last Updated"])
            for i, k in enumerate(kpis, 2):
                _data_row(ws_kpi, [
                    k.get("name"), k.get("current_value"), k.get("target_value"),
                    k.get("unit"), k.get("status"), k.get("trend"), k.get("last_updated")
                ], row=i, alt=(i % 2 == 0))
                # Colour the status cell
                status_cell = ws_kpi.cell(row=i, column=5)
                color = STATUS_COLORS.get(str(k.get("status") or "").lower(), "BFBFBF")
                status_cell.fill = PatternFill("solid", fgColor=color)
                status_cell.font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
            _auto_width(ws_kpi)

            # Sheet 3: Milestones
            ws_ms = wb.create_sheet("Milestones")
            _hdr_row(ws_ms, ["Milestone", "Due Date", "Completed Date", "Owner", "Status", "Description"])
            for i, m in enumerate(milestones, 2):
                _data_row(ws_ms, [
                    m.get("name"), m.get("due_date"), m.get("completed_date"),
                    m.get("owner"), m.get("status"), m.get("description")
                ], row=i, alt=(i % 2 == 0))
            _auto_width(ws_ms)

            # Sheet 4: Risks
            ws_risk = wb.create_sheet("Risks")
            _hdr_row(ws_risk, ["Risk", "Severity", "Impact", "Likelihood", "Status", "Mitigation", "Owner"])
            for i, r in enumerate(risks, 2):
                _data_row(ws_risk, [
                    r.get("title"), r.get("severity"), r.get("impact"),
                    r.get("likelihood"), r.get("status"), r.get("mitigation_plan"), r.get("owner")
                ], row=i, alt=(i % 2 == 0))
            _auto_width(ws_risk)

            # Sheet 5: Budget
            ws_bud = wb.create_sheet("Budget")
            _hdr_row(ws_bud, ["Category", "Allocated ($)", "Spent ($)", "Remaining ($)", "Notes"])
            total_alloc = total_spent = total_remain = 0
            for i, b in enumerate(budget, 2):
                alloc   = b.get("allocated") or 0
                spent   = b.get("spent") or 0
                remain  = b.get("remaining") or (alloc - spent)
                total_alloc  += alloc
                total_spent  += spent
                total_remain += remain
                _data_row(ws_bud, [b.get("category"), alloc, spent, remain, b.get("notes")],
                          row=i, alt=(i % 2 == 0))
                # Format as currency
                for col in [2, 3, 4]:
                    ws_bud.cell(row=i, column=col).number_format = '#,##0.00'
            # Totals row
            total_row = len(budget) + 2
            for c, v in enumerate(["TOTAL", total_alloc, total_spent, total_remain, ""], 1):
                cell = ws_bud.cell(row=total_row, column=c, value=v)
                cell.font = Font(bold=True, name="Calibri", size=10)
                if c in (2, 3, 4):
                    cell.number_format = '#,##0.00'
            _auto_width(ws_bud)

            # Sheet 6: Action Queue
            ws_act = wb.create_sheet("Actions")
            _hdr_row(ws_act, ["Title", "Category", "Priority", "Status", "Department", "Created", "Reviewed By"])
            for i, a in enumerate(actions, 2):
                _data_row(ws_act, [
                    a.get("title"), a.get("category"), a.get("priority"),
                    a.get("status"), a.get("agent_dept"),
                    str(a.get("created_at") or "")[:10], a.get("reviewed_by")
                ], row=i, alt=(i % 2 == 0))
            _auto_width(ws_act)

            # ── Save ───────────────────────────────────────────────────────
            out_dir  = self._output_dir(context)
            filename = f"export_{uuid.uuid4().hex[:8]}.xlsx"
            out_path = os.path.join(out_dir, filename)
            wb.save(out_path)

            n_sheets = len(wb.sheetnames)
            return SkillOutput(
                skill_id    = self.skill_id,
                dept_id     = self.dept_id,
                title       = title,
                file_format = self.output_format,
                file_path   = out_path,
                preview     = (
                    f"Excel workbook: {n_sheets} sheets — "
                    f"{len(kpis)} KPIs, {len(milestones)} milestones, "
                    f"{len(risks)} risks, {len(actions)} actions."
                ),
                pages       = n_sheets,
            )

        except Exception as e:
            logger.error(f"[ExportSkill] Failed: {e}", exc_info=True)
            return SkillOutput(
                skill_id    = self.skill_id,
                dept_id     = self.dept_id,
                title       = title,
                file_format = self.output_format,
                error       = str(e),
            )


# ── Self-registration ─────────────────────────────────────────────────────────

def _register(registry) -> None:
    registry.register(ExportSkill())
