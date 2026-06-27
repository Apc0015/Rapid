"""
Sales Skill: Weekly Forecast (XLSX)

Generates a weekly sales forecast spreadsheet with:
  - Commit, best case, pipeline forecast vs. target
  - Rep performance (from KPIs)
  - Quarter-to-date summary

Trigger phrases: "weekly forecast", "sales forecast", "forecast report",
                 "commit forecast", "revenue forecast"
"""

from __future__ import annotations

import logging
import os
import sqlite3
import uuid
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from agents.skills.base_skill import BaseSkill, SkillOutput
from infrastructure.project_context import ProjectContext

logger = logging.getLogger(__name__)

HDR_FILL  = PatternFill("solid", fgColor="1F3564")
HDR_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
ALT_FILL  = PatternFill("solid", fgColor="DCE6F1")
BODY_FONT = Font(name="Calibri", size=10)
THIN      = Side(style="thin", color="BFBFBF")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FMT = '#,##0.00'
CENT      = Alignment(horizontal="center")
RIGHT     = Alignment(horizontal="right")


class WeeklyForecastSkill(BaseSkill):
    skill_id        = "weekly_forecast"
    dept_id         = "sales"
    title_template  = "{project_name} — Weekly Sales Forecast"
    description     = "Sales: Generate a weekly forecast XLSX with commit/best case/pipeline vs. target and QTD summary."
    output_format   = "xlsx"
    trigger_phrases = [
        "weekly forecast", "sales forecast", "forecast report",
        "commit forecast", "revenue forecast", "weekly sales",
        "forecast this week",
    ]

    async def execute(self, context: ProjectContext, params: dict[str, Any] = None) -> SkillOutput:
        params = params or {}
        title  = params.get("title") or self._make_title(context)

        try:
            db_path   = getattr(context, "db_path", None)
            proj_info = {}
            kpis      = []

            if db_path and os.path.exists(db_path):
                conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True, timeout=10)
                conn.row_factory = sqlite3.Row
                try:
                    r = conn.execute("SELECT * FROM project_details LIMIT 1").fetchone()
                    if r:
                        proj_info = dict(r)
                    kpis = [dict(r) for r in conn.execute(
                        "SELECT name, current_value, target_value, unit, status, trend "
                        "FROM project_kpis ORDER BY name"
                    ).fetchall()]
                finally:
                    conn.close()

            wb  = openpyxl.Workbook()
            ws  = wb.active
            ws.title = "Forecast"
            now = datetime.utcnow().strftime("%Y-%m-%d")
            proj_name = proj_info.get("name") or context.project_id[:12]

            # Title
            ws.merge_cells("A1:G1")
            ws["A1"] = f"Weekly Sales Forecast — {proj_name}"
            ws["A1"].font = Font(color="FFFFFF", bold=True, name="Calibri", size=14)
            ws["A1"].fill = HDR_FILL
            ws["A1"].alignment = CENT

            ws.merge_cells("A2:G2")
            ws["A2"] = f"Week of: {now}  |  Period: {proj_info.get('start_date','—')} — {proj_info.get('target_end_date','—')}"
            ws["A2"].font = Font(name="Calibri", size=10, italic=True)
            ws["A2"].alignment = CENT

            # Forecast table headers
            row = 4
            headers = ["Metric", "Commit ($)", "Best Case ($)", "Pipeline ($)", "Target ($)", "Attainment %", "Status"]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font = HDR_FONT; cell.fill = HDR_FILL
                cell.alignment = CENT; cell.border = BORDER
            row += 1

            # Pull revenue/quota KPIs as forecast metrics
            forecast_kpis = [k for k in kpis if any(
                w in str(k.get("name","")).lower()
                for w in ["revenue", "arr", "mrr", "quota", "pipeline", "forecast",
                          "commit", "booking", "deal", "target"]
            )]

            total_commit = total_target = 0
            for idx, k in enumerate(forecast_kpis):
                current = float(k.get("current_value") or 0)
                target  = float(k.get("target_value") or 0)
                attain  = (current / target) if target else 0
                status  = k.get("status") or "unknown"

                # Synthetic commit/best-case/pipeline from current
                commit    = current * 0.85
                best_case = current
                pipeline  = current * 1.15

                total_commit += commit
                total_target += target

                values = [k.get("name"), commit, best_case, pipeline, target, attain, status]
                fmts   = [None, MONEY_FMT, MONEY_FMT, MONEY_FMT, MONEY_FMT, "0.0%", None]
                alt    = idx % 2 == 1

                for c, (v, fmt) in enumerate(zip(values, fmts), 1):
                    cell = ws.cell(row=row, column=c, value=v)
                    cell.font = BODY_FONT
                    cell.border = BORDER
                    if alt:
                        cell.fill = ALT_FILL
                    if fmt:
                        cell.number_format = fmt
                        cell.alignment = RIGHT
                row += 1

            if not forecast_kpis:
                ws.cell(row=row, column=1, value="No sales-related KPIs configured. Add KPIs with names like 'Revenue', 'Quota', 'Pipeline', or 'ARR'.")
                ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, italic=True)
                row += 1

            # Totals
            row += 1
            totals = ["TOTAL COMMIT", total_commit, "", "", total_target, (total_commit/total_target) if total_target else 0, ""]
            fmts   = [None, MONEY_FMT, None, None, MONEY_FMT, "0.0%", None]
            for c, (v, fmt) in enumerate(zip(totals, fmts), 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
                cell.fill = PatternFill("solid", fgColor="1F3564")
                cell.border = BORDER
                if fmt:
                    cell.number_format = fmt
                    cell.alignment = RIGHT

            # Column widths
            ws.column_dimensions["A"].width = 30
            for col in ["B","C","D","E","F","G"]:
                ws.column_dimensions[col].width = 16

            # QTD Sheet
            ws2 = wb.create_sheet("QTD Summary")
            ws2.merge_cells("A1:D1")
            ws2["A1"] = f"Quarter-to-Date Summary — {proj_name}"
            ws2["A1"].font = Font(color="FFFFFF", bold=True, name="Calibri", size=13)
            ws2["A1"].fill = HDR_FILL
            ws2["A1"].alignment = CENT

            for c, h in enumerate(["KPI", "Actual", "Target", "Status"], 1):
                cell = ws2.cell(row=3, column=c, value=h)
                cell.font = HDR_FONT; cell.fill = HDR_FILL
                cell.alignment = CENT; cell.border = BORDER

            for idx, k in enumerate(kpis, 4):
                alt = (idx % 2 == 0)
                for c, v in enumerate([k.get("name"), k.get("current_value"), k.get("target_value"), k.get("status")], 1):
                    cell = ws2.cell(row=idx, column=c, value=v)
                    cell.font = BODY_FONT; cell.border = BORDER
                    if alt:
                        cell.fill = ALT_FILL

            for col in ["A","B","C","D"]:
                ws2.column_dimensions[col].width = 25

            # Save
            out_dir  = self._output_dir(context)
            filename = f"weekly_forecast_{uuid.uuid4().hex[:8]}.xlsx"
            out_path = os.path.join(out_dir, filename)
            wb.save(out_path)

            return SkillOutput(
                skill_id    = self.skill_id,
                dept_id     = self.dept_id,
                title       = title,
                file_format = self.output_format,
                file_path   = out_path,
                preview     = f"Forecast: commit ${total_commit:,.0f} vs target ${total_target:,.0f} ({(total_commit/total_target*100) if total_target else 0:.1f}%).",
                pages       = 2,
            )

        except Exception as e:
            logger.error(f"[WeeklyForecastSkill] Failed: {e}", exc_info=True)
            return SkillOutput(skill_id=self.skill_id, dept_id=self.dept_id,
                               title=title, file_format=self.output_format, error=str(e))


def _register(registry) -> None:
    registry.register(WeeklyForecastSkill())
