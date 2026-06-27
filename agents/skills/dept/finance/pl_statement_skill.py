"""
Finance Skill: P&L Statement (XLSX)

Generates a structured Profit & Loss statement spreadsheet from project
financial data. For projects with revenue and cost tracking.

Trigger phrases: "p&l", "profit and loss", "income statement", "pl statement"
"""

from __future__ import annotations

import logging
import os
import sqlite3
import uuid
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from agents.skills.base_skill import BaseSkill, SkillOutput
from infrastructure.project_context import ProjectContext

logger = logging.getLogger(__name__)

HDR_FILL   = PatternFill("solid", fgColor="1F3564")
HDR_FONT   = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
SUB_FILL   = PatternFill("solid", fgColor="2E75B6")
SUB_FONT   = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
BOLD_FONT  = Font(bold=True, name="Calibri", size=10)
BODY_FONT  = Font(name="Calibri", size=10)
THIN       = Side(style="thin", color="BFBFBF")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FMT  = '#,##0.00'
CENT       = Alignment(horizontal="center")
RIGHT      = Alignment(horizontal="right")


class PLStatementSkill(BaseSkill):
    skill_id        = "pl_statement"
    dept_id         = "finance"
    title_template  = "{project_name} — P&L Statement"
    description     = "Finance: Generate a Profit & Loss statement XLSX with revenue, costs, gross margin, and net income."
    output_format   = "xlsx"
    trigger_phrases = [
        "p&l", "profit and loss", "income statement", "pl statement",
        "p and l", "pnl", "revenue and costs", "financial statement",
    ]

    async def execute(self, context: ProjectContext, params: dict[str, Any] = None) -> SkillOutput:
        params = params or {}
        title  = params.get("title") or self._make_title(context)

        try:
            db_path   = getattr(context, "db_path", None)
            proj_info = {}
            kpis      = []
            budget    = []

            if db_path and os.path.exists(db_path):
                conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True, timeout=10)
                conn.row_factory = sqlite3.Row
                try:
                    r = conn.execute("SELECT * FROM project_details LIMIT 1").fetchone()
                    if r:
                        proj_info = dict(r)

                    kpis = [dict(r) for r in conn.execute(
                        "SELECT name, current_value, target_value, unit, status FROM project_kpis"
                    ).fetchall()]

                    budget = [dict(r) for r in conn.execute(
                        "SELECT category, allocated, spent FROM project_budget_lines"
                    ).fetchall()]
                finally:
                    conn.close()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "P&L Statement"

            now       = datetime.utcnow().strftime("%Y-%m-%d")
            proj_name = proj_info.get("name") or context.project_id[:12]

            # Header block
            ws.merge_cells("A1:E1")
            ws["A1"] = f"Profit & Loss Statement — {proj_name}"
            ws["A1"].font      = Font(color="FFFFFF", bold=True, name="Calibri", size=14)
            ws["A1"].fill      = HDR_FILL
            ws["A1"].alignment = CENT

            ws.merge_cells("A2:E2")
            ws["A2"] = f"Period: {proj_info.get('start_date','—')} to {proj_info.get('target_end_date','—')}  |  Generated: {now}"
            ws["A2"].font      = Font(name="Calibri", size=10, italic=True)
            ws["A2"].alignment = CENT

            row = 4
            # Column headers
            for c, h in enumerate(["Category", "Budget ($)", "Actual ($)", "Variance ($)", "Variance %"], 1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font      = HDR_FONT
                cell.fill      = HDR_FILL
                cell.alignment = CENT
                cell.border    = BORDER
            row += 1

            def _write_row(label, budget_v, actual_v, bold=False, fill=None):
                nonlocal row
                var_v   = budget_v - actual_v
                var_pct = (var_v / budget_v * 100) if budget_v else 0
                values  = [label, budget_v, actual_v, var_v, var_pct / 100]
                fmts    = [None, MONEY_FMT, MONEY_FMT, MONEY_FMT, "0.0%"]
                font    = Font(bold=bold, name="Calibri", size=10)
                for c, (v, fmt) in enumerate(zip(values, fmts), 1):
                    cell = ws.cell(row=row, column=c, value=v)
                    cell.font   = font
                    cell.border = BORDER
                    if fmt:
                        cell.number_format = fmt
                        cell.alignment = RIGHT
                    if fill:
                        cell.fill = fill
                row += 1
                return var_v

            # Revenue section
            ws.cell(row=row, column=1, value="REVENUE").font = Font(bold=True, name="Calibri", size=11, color="FFFFFF")
            ws.cell(row=row, column=1).fill = SUB_FILL
            ws.merge_cells(f"A{row}:E{row}")
            row += 1

            revenue_kpis = [k for k in kpis if any(
                w in k.get("name","").lower() for w in ["revenue", "sales", "income", "arr", "mrr"]
            )]
            total_rev_bud = total_rev_act = 0
            for k in revenue_kpis:
                bud = float(k.get("target_value") or 0)
                act = float(k.get("current_value") or 0)
                _write_row(k.get("name"), bud, act)
                total_rev_bud += bud
                total_rev_act += act
            if not revenue_kpis:
                # Fall back to budget "revenue" lines
                for b in budget:
                    if "revenue" in str(b.get("category","")).lower():
                        bud = b.get("allocated") or 0
                        act = b.get("spent") or 0
                        _write_row(b.get("category"), bud, act)
                        total_rev_bud += bud
                        total_rev_act += act

            _write_row("Total Revenue", total_rev_bud, total_rev_act,
                       bold=True, fill=PatternFill("solid", fgColor="D9E1F2"))
            row += 1  # spacer

            # Costs section
            ws.cell(row=row, column=1, value="COSTS & EXPENSES").font = Font(bold=True, name="Calibri", size=11, color="FFFFFF")
            ws.cell(row=row, column=1).fill = SUB_FILL
            ws.merge_cells(f"A{row}:E{row}")
            row += 1

            total_cost_bud = total_cost_act = 0
            for b in budget:
                if "revenue" not in str(b.get("category","")).lower():
                    bud = b.get("allocated") or 0
                    act = b.get("spent") or 0
                    _write_row(b.get("category"), bud, act)
                    total_cost_bud += bud
                    total_cost_act += act

            _write_row("Total Costs", total_cost_bud, total_cost_act,
                       bold=True, fill=PatternFill("solid", fgColor="D9E1F2"))
            row += 1  # spacer

            # Net Income row
            net_bud = total_rev_bud - total_cost_bud
            net_act = total_rev_act - total_cost_act
            _write_row("NET INCOME / (LOSS)", net_bud, net_act,
                       bold=True, fill=PatternFill("solid", fgColor="1F3564"))
            # White font for net income row
            for c in range(1, 6):
                ws.cell(row=row-1, column=c).font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)

            # Column widths
            ws.column_dimensions["A"].width = 35
            for col in ["B","C","D","E"]:
                ws.column_dimensions[col].width = 18

            # Save
            out_dir  = self._output_dir(context)
            filename = f"pl_statement_{uuid.uuid4().hex[:8]}.xlsx"
            out_path = os.path.join(out_dir, filename)
            wb.save(out_path)

            return SkillOutput(
                skill_id    = self.skill_id,
                dept_id     = self.dept_id,
                title       = title,
                file_format = self.output_format,
                file_path   = out_path,
                preview     = (
                    f"P&L: Revenue budget ${total_rev_bud:,.0f} | "
                    f"Costs budget ${total_cost_bud:,.0f} | "
                    f"Net ${net_bud:,.0f}"
                ),
                pages       = 1,
            )

        except Exception as e:
            logger.error(f"[PLStatementSkill] Failed: {e}", exc_info=True)
            return SkillOutput(skill_id=self.skill_id, dept_id=self.dept_id,
                               title=title, file_format=self.output_format, error=str(e))


def _register(registry) -> None:
    registry.register(PLStatementSkill())
